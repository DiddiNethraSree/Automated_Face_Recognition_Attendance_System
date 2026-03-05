from flask import Flask, render_template, request, redirect, session, send_file, jsonify
import sqlite3
import database
from datetime import datetime, timedelta
import database as dbmod
from database import PERIODS
import io
from collections import OrderedDict

# Map HOD username prefix → department code used in student IDs
# Any <prefix>hod username is valid; prefix must match a key below.
HOD_DEPT_MAP = {
    'cs': 'CSE',
    'cse': 'CSE',
    'ece': 'ECE',
    'ec': 'ECE',
    'eee': 'EEE',
    'ee': 'EE',
    'mech': 'MECH',
    'me': 'MECH',
    'civil': 'CIVIL',
    'cl': 'CIVIL',
    'ce': 'CIVIL',
    'aiml': 'AIML',
    'ai': 'AIML',
    'it': 'IT',
    'ds': 'DS',
    'csd': 'CSD',
    'csm': 'CSM',
    'csbs': 'CSBS',
    'cyber': 'CYBER',
    'cy': 'CYBER',
    'cs(cyber)': 'CYBER',
    'eie': 'EIE',
    'ei': 'EIE',
    'iem': 'IEM',
    'aero': 'AERO',
    'ae': 'AERO',
    'biotech': 'BIOTECH',
    'bt': 'BIOTECH',
    'chem': 'CHEM',
    'mines': 'MINES',
    'mining': 'MINES',
    'pharma': 'PHARMA',
    'csit': 'CSIT',
    'agri': 'AGRI',
    'pete': 'PETE',
}

# All B.Tech specializations offered at Bapatla Engineering College
ALL_BRANCHES = [
    ('CIVIL',  'civil',  'Civil Engineering'),
    ('CYBER',  'cyber',  'Cyber Security'),
    ('CSE',    'cs',     'Computer Science & Engineering'),
    ('CSM',    'csm',    'CSE - AI & ML'),
    ('DS',     'ds',     'Data Science'),
    ('ECE',    'ece',    'Electronics & Communication Engineering'),
    ('EEE',    'eee',    'Electrical & Electronics Engineering'),
    ('EIE',    'eie',    'Electronics & Instrumentation Engineering'),
    ('IT',     'it',     'Information Technology'),
    ('MECH',   'mech',   'Mechanical Engineering'),
]

def get_hod_department(hod_user_id):
    """Extract department code from HOD username. e.g. 'cshod' → 'CSE', 'ithod' → 'IT'."""
    uid = hod_user_id.lower().strip()
    if uid.endswith('hod'):
        prefix = uid[:-3]  # strip trailing 'hod'
    else:
        prefix = uid
    return HOD_DEPT_MAP.get(prefix)

app = Flask(__name__)
app.secret_key = "attendance_secret"

# Initialize DB on startup
database.init_db()

DB = "attendance.db"

def get_db():
    return sqlite3.connect(DB)

# ---------- LOGIN ----------
@app.route("/", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        user_id = request.form.get("user_id")
        password = request.form.get("password")
        role = request.form.get("role")

        con = sqlite3.connect(DB)
        cur = con.cursor()

        cur.execute("SELECT user_id, role FROM users WHERE lower(user_id)=lower(?) AND password=?", (user_id, password))
        row = cur.fetchone()
        con.close()

        if row:
            session["user_id"] = row[0]
            session["role"] = row[1]

            if row[1] == "student":
                return redirect("/student")
            elif row[1] == "hod":
                return redirect("/hod")

        return render_template("login.html", error="Invalid credentials")

    return render_template("login.html")


# ---------- LOGOUT ----------
@app.route("/logout")
def logout():
    session.clear()
    return redirect("/")


# ---------- FORGOT PASSWORD ROUTES ----------

@app.route("/forgot-password", methods=["GET", "POST"])
def forgot_password():
    if request.method == "POST":
        user_id = request.form.get("reg_no")
        dob = request.form.get("dob")

        con = get_db()
        cur = con.cursor()
        cur.execute("SELECT user_id FROM users WHERE lower(user_id)=lower(?) AND dob=? AND role='student'", (user_id, dob))
        user = cur.fetchone()
        con.close()

        if user:
            session["reset_user_id"] = user_id
            return redirect("/reset-password")
        else:
            return render_template("forget_password_student.html", error="Invalid Register Number or DOB")

    return render_template("forget_password_student.html")


@app.route("/reset-password", methods=["GET", "POST"])
def reset_password():
    if "reset_user_id" not in session:
        return redirect("/forgot-password")

    if request.method == "POST":
        new_password = request.form.get("new_password")
        confirm_password = request.form.get("confirm_password")

        if new_password != confirm_password:
            return render_template("rest_password_student.html", error="Passwords do not match")

        user_id = session["reset_user_id"]
        con = get_db()
        cur = con.cursor()
        cur.execute("UPDATE users SET password=? WHERE user_id=?", (new_password, user_id))
        con.commit()
        con.close()

        session.pop("reset_user_id", None)
        return render_template("reset_success.html")

    return render_template("rest_password_student.html")


# ---------- HOD DASHBOARD ----------
@app.route("/hod")
def hod():
    if session.get("role") != "hod":
        return redirect("/")

    hod_uid = session["user_id"]
    hod_dept = get_hod_department(hod_uid)  # e.g. 'EE', 'CSE', or None for 'hod1'

    time_filter = request.args.get("filter", "all")
    year_filter = request.args.get("year")
    branch_filter = request.args.get("branch")
    section_filter = request.args.get("section")
    q_filter = request.args.get("q")
    date_filter = request.args.get("date")
    con = get_db()
    cur = con.cursor()
    cur.execute("SELECT user_id FROM users WHERE role='student'")
    students_seed = [r[0] for r in cur.fetchall()]
    today_str = dbmod.now_ist().strftime("%Y-%m-%d")
    if students_seed and dbmod.is_working_day(today_str):
        database.init_today(students_seed)
        database.init_today_periods(students_seed)

    # Backfill any missing past working days for all students
    database.backfill_all_students()

    # --- Build common WHERE clause with department auto-filter ---
    def build_where(table_alias):
        clauses = []
        params = []
        # Auto-filter by HOD department: use branch column when set, fallback to student_id substring
        if hod_dept:
            clauses.append(f"(lower(u.branch) = lower(?) OR (u.branch IS NULL AND lower({table_alias}.student_id) LIKE '%' || lower(?) || '%'))")
            params.extend([hod_dept, hod_dept])
        if time_filter == 'today':
            clauses.append(f"{table_alias}.date = ?")
            params.append(today_str)
        elif time_filter != 'all':
            days = {'7days': 7, '15days': 15, '30days': 30}.get(time_filter, 7)
            start_date = (dbmod.now_ist() - timedelta(days=days)).strftime("%Y-%m-%d")
            clauses.append(f"{table_alias}.date >= ?")
            params.append(start_date)
        if year_filter:
            clauses.append(f"(u.year = ? OR (CASE WHEN substr({table_alias}.student_id,1,2) IN ('21','22') THEN '4' WHEN substr({table_alias}.student_id,1,2) = '23' THEN '3' WHEN substr({table_alias}.student_id,1,2) = '24' THEN '2' WHEN substr({table_alias}.student_id,1,2) = '25' THEN '1' ELSE NULL END) = ?)")
            params.extend([year_filter, year_filter])
        if branch_filter:
            clauses.append(f"(lower(u.branch) = lower(?) OR lower({table_alias}.student_id) LIKE '%' || lower(?) || '%')")
            params.extend([branch_filter, branch_filter])
        if section_filter:
            clauses.append("lower(u.section) = lower(?)")
            params.append(section_filter)
        if q_filter:
            clauses.append(f"lower({table_alias}.student_id) LIKE '%' || lower(?) || '%'")
            params.append(q_filter)
        if date_filter:
            clauses.append(f"{table_alias}.date = ?")
            params.append(date_filter)
        return clauses, params

    # Period-wise attendance percentage
    w1, p1 = build_where('pa')
    query = "SELECT pa.student_id, ROUND(SUM(pa.value)*100.0/COUNT(*),2) AS percent FROM period_attendance pa LEFT JOIN users u ON u.user_id = pa.student_id"
    if w1:
        query += " WHERE " + " AND ".join(w1)
    query += " GROUP BY pa.student_id"
    cur.execute(query, p1)
    data = cur.fetchall()
    
    # Period-wise master list
    w2, p2 = build_where('pa')
    list_query = "SELECT pa.student_id, pa.date, pa.period, pa.status, pa.value, pa.first_seen_time FROM period_attendance pa LEFT JOIN users u ON u.user_id = pa.student_id"
    if w2:
        list_query += " WHERE " + " AND ".join(w2)
    list_query += " ORDER BY pa.date DESC, pa.student_id ASC, pa.period ASC"
    cur.execute(list_query, p2)
    master_list = cur.fetchall()

    cur.execute("SELECT student_id, ROUND(SUM(value)*100.0/COUNT(*),2) FROM period_attendance GROUP BY student_id")
    overall_percentages = {s: float(p) for s, p in cur.fetchall()}
    filtered_percentages = {d[0]: float(d[1]) for d in data}

    con.close()

    eligible = [d for d in data if d[1] >= 75]
    condonation = [d for d in data if 65 <= d[1] < 75]
    detained = [d for d in data if d[1] < 65]

    # Build period schedule for template
    period_schedule = {k: f"{v[0]:02d}:{v[1]:02d}\u2013{v[2]:02d}:{v[3]:02d}" for k, v in PERIODS.items()}

    # Group master_list by date for day-wise accordion view
    days_data = OrderedDict()
    for row in master_list:
        sid, dt, period, status, value, time_in = row
        if dt not in days_data:
            days_data[dt] = {'students': {}}
        if sid not in days_data[dt]['students']:
            days_data[dt]['students'][sid] = {'periods': {}}
        days_data[dt]['students'][sid]['periods'][period] = {
            'status': status, 'value': float(value), 'time': time_in or ""
        }

    return render_template(
        "hod_dashboard.html",
        eligible=eligible,
        condonation=condonation,
        detained=detained,
        master_list=master_list,
        days_data=days_data,
        overall_percentages=overall_percentages,
        filtered_percentages=filtered_percentages,
        current_filter=time_filter,
        is_today_working=dbmod.is_working_day(today_str),
        period_schedule=period_schedule,
        current_year=year_filter or "",
        current_branch=branch_filter or "",
        current_section=section_filter or "",
        current_q=q_filter or "",
        hod_dept=hod_dept or 'ALL',
        today_date=today_str
    )


# ---------- STUDENT DASHBOARD ----------
@app.route("/student")
def student():
    if session.get("role") != "student":
        return redirect("/")

    sid = session["user_id"]
    today_str = dbmod.now_ist().strftime("%Y-%m-%d")
    if dbmod.is_working_day(today_str):
        database.init_today([sid])
        database.init_today_periods([sid])

    # Backfill any missing past working days for this student
    database.backfill_student_periods(sid)

    con = get_db()
    cur = con.cursor()

    # Period-wise records
    cur.execute(
        "SELECT date, period, status, value, first_seen_time FROM period_attendance WHERE student_id=? ORDER BY date DESC, period ASC",
        (sid,)
    )
    period_rows = cur.fetchall()

    # Today's period breakdown
    cur.execute(
        "SELECT period, status, value, first_seen_time FROM period_attendance WHERE student_id=? AND date=? ORDER BY period",
        (sid, today_str)
    )
    today_periods = cur.fetchall()

    # Period-wise overall stats
    total_periods = len(period_rows)
    total_value = sum(r[3] for r in period_rows)
    percent = round((total_value / total_periods) * 100, 2) if total_periods else 0

    # Count stats
    present_count = sum(1 for r in period_rows if r[2] == 'present')
    absent_count = sum(1 for r in period_rows if r[2] == 'absent')

    con.close()

    period_schedule = {k: f"{v[0]:02d}:{v[1]:02d}–{v[2]:02d}:{v[3]:02d}" for k, v in PERIODS.items()}

    # Group period records by date for day-wise view
    from collections import OrderedDict
    days_attendance = OrderedDict()
    for r in period_rows:  # already sorted date DESC, period ASC
        date_str_key = r[0]
        if date_str_key not in days_attendance:
            days_attendance[date_str_key] = []
        days_attendance[date_str_key].append({
            'period': r[1],
            'status': r[2],
            'value': r[3],
            'time_in': r[4] if r[4] else '--'
        })

    return render_template(
        "student_dashboard.html",
        student_id=sid,
        percent=percent,
        period_records=period_rows,
        today_periods=today_periods,
        total_periods=total_periods,
        present_count=present_count,
        absent_count=absent_count,
        is_today_working=dbmod.is_working_day(today_str),
        period_schedule=period_schedule,
        days_attendance=days_attendance,
        today_date=today_str
    )

@app.route("/api/hod-data")
def api_hod_data():
    if session.get("role") != "hod":
        return {"error": "unauthorized"}, 401
    hod_uid = session["user_id"]
    hod_dept = get_hod_department(hod_uid)

    time_filter = request.args.get("filter", "all")
    year_filter = request.args.get("year")
    branch_filter = request.args.get("branch")
    section_filter = request.args.get("section")
    section_filter = request.args.get("section")
    q_filter = request.args.get("q")
    date_filter = request.args.get("date")
    derive_toggle = request.args.get("derive") == "1"

    con = get_db()
    cur = con.cursor()

    def build_where(table_alias):
        clauses = []
        params = []
        if hod_dept:
            clauses.append(f"(lower(u.branch) = lower(?) OR (u.branch IS NULL AND lower({table_alias}.student_id) LIKE '%' || lower(?) || '%'))")
            params.extend([hod_dept, hod_dept])
        if time_filter == 'today':
            today_api = dbmod.now_ist().strftime("%Y-%m-%d")
            clauses.append(f"{table_alias}.date = ?")
            params.append(today_api)
        elif time_filter != 'all':
            days = {'7days': 7, '15days': 15, '30days': 30}.get(time_filter, 7)
            start_date = (dbmod.now_ist() - timedelta(days=days)).strftime("%Y-%m-%d")
            clauses.append(f"{table_alias}.date >= ?")
            params.append(start_date)
        if year_filter:
            if derive_toggle:
                clauses.append(f"(u.year = ? OR (CASE WHEN substr({table_alias}.student_id,1,1)='L' AND substr({table_alias}.student_id,2,2)='23' THEN '4' WHEN substr({table_alias}.student_id,1,2)='22' THEN '2' WHEN substr({table_alias}.student_id,1,2)='21' THEN '4' WHEN substr({table_alias}.student_id,1,2)='23' THEN '3' WHEN substr({table_alias}.student_id,1,2)='24' THEN '2' WHEN substr({table_alias}.student_id,1,2)='25' THEN '1' WHEN substr({table_alias}.student_id,1,2)='20' THEN '5' ELSE NULL END) = ?)")
            else:
                clauses.append(f"(u.year = ? OR (CASE WHEN substr({table_alias}.student_id,1,2) IN ('21','22') THEN '4' WHEN substr({table_alias}.student_id,1,2) = '23' THEN '3' WHEN substr({table_alias}.student_id,1,2) = '24' THEN '2' WHEN substr({table_alias}.student_id,1,2) = '25' THEN '1' ELSE NULL END) = ?)")
            params.extend([year_filter, year_filter])
        if branch_filter:
            clauses.append(f"(lower(u.branch) = lower(?) OR lower({table_alias}.student_id) LIKE '%' || lower(?) || '%')")
            params.extend([branch_filter, branch_filter])
        if section_filter:
            clauses.append("lower(u.section) = lower(?)")
            params.append(section_filter)
        if q_filter:
            clauses.append(f"lower({table_alias}.student_id) LIKE '%' || lower(?) || '%'")
            params.append(q_filter)
        if date_filter:
            clauses.append(f"{table_alias}.date = ?")
            params.append(date_filter)
        return clauses, params

    w1, p1 = build_where('pa')
    query = "SELECT pa.student_id, ROUND(SUM(pa.value)*100.0/COUNT(*),2) AS percent FROM period_attendance pa LEFT JOIN users u ON u.user_id = pa.student_id"
    if w1:
        query += " WHERE " + " AND ".join(w1)
    query += " GROUP BY pa.student_id"
    cur.execute(query, p1)
    data = cur.fetchall()

    filtered_percentages = {s: float(p) for s, p in data}
    eligible = [ {"id": s, "percent": float(p)} for s, p in data if p >= 75 ]
    condonation = [ {"id": s, "percent": float(p)} for s, p in data if 65 <= p < 75 ]
    detained = [ {"id": s, "percent": float(p)} for s, p in data if p < 65 ]

    cur.execute("SELECT student_id, ROUND(SUM(value)*100.0/COUNT(*),2) FROM period_attendance GROUP BY student_id")
    overall_percentages = {s: float(p) for s, p in cur.fetchall()}

    w2, p2 = build_where('pa')
    list_query = "SELECT pa.student_id, pa.date, pa.period, pa.status, pa.value, pa.first_seen_time FROM period_attendance pa LEFT JOIN users u ON u.user_id = pa.student_id"
    if w2:
        list_query += " WHERE " + " AND ".join(w2)
    list_query += " ORDER BY pa.date DESC, pa.student_id ASC, pa.period ASC"
    cur.execute(list_query, p2)
    
    from collections import OrderedDict
    days_data = OrderedDict()
    for row in cur.fetchall():
        sid, dt, period, status, value, time_in = row
        if dt not in days_data:
            days_data[dt] = {'students': {}}
        
        if sid not in days_data[dt]['students']:
            days_data[dt]['students'][sid] = {'periods': {}}
        
        days_data[dt]['students'][sid]['periods'][period] = {
            'status': status, 'value': float(value), 'time': time_in or ""
        }

    con.close()

    return {
        "eligible": eligible,
        "condonation": condonation,
        "detained": detained,
        "days_data": days_data,
        "overall_percentages": overall_percentages,
        "filtered_percentages": filtered_percentages
    }
@app.route("/api/check-username")
def api_check_username():
    user_id = request.args.get("user_id","").lower()
    con = get_db()
    cur = con.cursor()
    cur.execute("SELECT 1 FROM users WHERE lower(user_id)=?", (user_id,))
    exists = cur.fetchone() is not None
    con.close()
    return {"exists": exists}
@app.route("/hod/manual-attendance", methods=["GET","POST"])
def hod_manual_attendance():
    if session.get("role") != "hod":
        return redirect("/")
    if request.method == "POST":
        reg_no = request.form.get("reg_no")
        date_str = request.form.get("date")
        time_str = request.form.get("time")
        year = request.form.get("year")
        period_val = request.form.get("period", "all")  # 1-6 or 'all'
        status_val = request.form.get("status", "present")  # present, late, od
        # Use a connection with a timeout to avoid 'database is locked'
        con = sqlite3.connect(DB, timeout=10)
        cur = con.cursor()
        try:
            cur.execute("SELECT 1 FROM users WHERE lower(user_id)=lower(?)", (reg_no,))
            exists = cur.fetchone()
            if not exists:
                con.close()
                return render_template("hod_manual_attendance.html", error="Student not found", preset_id=reg_no, period_schedule=PERIODS)
            if year:
                cur.execute("UPDATE users SET year=? WHERE lower(user_id)=lower(?)", (year, reg_no))
            # Legacy attendance table
            cur.execute("INSERT OR REPLACE INTO attendance (student_id, date, first_seen_time, present) VALUES (?, ?, ?, 1)", (reg_no, date_str, time_str))
            # Period attendance table
            status_map = {'present': ('present', 1.0), 'od': ('present', 1.0)}
            st, val = status_map.get(status_val, ('present', 1.0))
            if period_val == 'all':
                periods_to_mark = list(PERIODS.keys())
            else:
                periods_to_mark = [int(period_val)]
            for p in periods_to_mark:
                cur.execute("INSERT OR REPLACE INTO period_attendance (student_id, date, period, status, value, first_seen_time) VALUES (?, ?, ?, ?, ?, ?)",
                            (reg_no, date_str, p, st, val, time_str))
            con.commit()
        except sqlite3.OperationalError as e:
            con.rollback()
            con.close()
            return render_template("hod_manual_attendance.html", error="Database is busy. Please retry in a moment.", preset_id=reg_no, period_schedule=PERIODS)
        con.close()
        return redirect("/hod")
    return render_template("hod_manual_attendance.html", period_schedule=PERIODS)

@app.route("/hod/signup", methods=["GET","POST"])
def hod_signup():
    import re
    branches = ALL_BRANCHES
    if request.method == "POST":
        branch_code = request.form.get("branch_code", "").strip()  # e.g. 'cs', 'it'
        password = request.form.get("password", "")
        # Build username automatically: e.g. cshod, ithod
        uid_lower = branch_code.lower() + "hod"

        ok = (len(password) >= 10 and re.search(r"[A-Z]", password) and re.search(r"[a-z]", password) and re.search(r"\d", password) and re.search(r"[^A-Za-z0-9]", password))
        if not ok:
            return render_template("hod_signup.html", error="Password must be 10+ chars with uppercase, lowercase, number, and special character.", branches=branches)
        if branch_code not in [b[1] for b in ALL_BRANCHES]:
            return render_template("hod_signup.html", error="Invalid branch selected.", branches=branches)
        hod_dept = HOD_DEPT_MAP.get(branch_code)
        con = get_db()
        cur = con.cursor()
        cur.execute("SELECT 1 FROM users WHERE lower(user_id)=lower(?)", (uid_lower,))
        if cur.fetchone():
            con.close()
            return render_template("hod_signup.html", error=f"HOD account '{uid_lower}' already exists. Please log in.", branches=branches)
        cur.execute("INSERT INTO users (user_id, password, role, dob, branch) VALUES (?, ?, 'hod', NULL, ?)", (uid_lower, password, hod_dept))
        con.commit()
        con.close()
        return redirect("/")
    return render_template("hod_signup.html", branches=branches)
@app.route("/student/signup", methods=["GET", "POST"])
def student_signup():
    import os, base64
    CLEAN_FACES_DIR = r"E:\dataset-attendance\clean_faces"
    if not os.path.exists(CLEAN_FACES_DIR):
        CLEAN_FACES_DIR = os.path.join(os.path.dirname(__file__), "clean_faces")
        os.makedirs(CLEAN_FACES_DIR, exist_ok=True)

    if request.method == "POST":
        reg_no = request.form.get("user_id")
        name = request.form.get("name")
        dob = request.form.get("dob")
        year = request.form.get("year")
        section = request.form.get("section")
        branch = request.form.get("branch")
        password = request.form.get("password") or "123"
        images_b64 = request.form.getlist("images[]")

        con = get_db()
        cur = con.cursor()
        # Ensure columns exist (backwards compatibility)
        cur.execute("PRAGMA table_info(users)")
        cols = [c[1] for c in cur.fetchall()]
        if "name" not in cols:
            cur.execute("ALTER TABLE users ADD COLUMN name TEXT")
        if "year" not in cols:
            cur.execute("ALTER TABLE users ADD COLUMN year TEXT")
        if "section" not in cols:
            cur.execute("ALTER TABLE users ADD COLUMN section TEXT")
        if "branch" not in cols:
            cur.execute("ALTER TABLE users ADD COLUMN branch TEXT")

        cur.execute("SELECT 1 FROM users WHERE lower(user_id)=lower(?)", (reg_no,))
        if cur.fetchone():
            con.close()
            return render_template("student_signup.html", error="User already exists. Please sign in.", preset_id=reg_no)

        import cv2, numpy as np, pickle, face_recognition
        face_is_duplicate = False
        duplicate_id = None
        try:
            with open("encodings.pickle", "rb") as f:
                data_enc = pickle.load(f)
            known_encs = np.array(data_enc["encodings"])
            known_names = np.array(data_enc["names"])
        except Exception:
            known_encs, known_names = [], []

        if len(known_encs) > 0 and len(images_b64) > 0:
            b64img = images_b64[0]
            try:
                header, data = b64img.split(',', 1) if ',' in b64img else ('', b64img)
                img_bytes = base64.b64decode(data)
                nparr = np.frombuffer(img_bytes, np.uint8)
                img = cv2.imdecode(nparr, cv2.IMREAD_COLOR)
                if img is not None:
                    rgb = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
                    boxes = face_recognition.face_locations(rgb, model="hog")
                    encs = face_recognition.face_encodings(rgb, boxes)
                    if encs:
                        distances = face_recognition.face_distance(known_encs, encs[0])
                        if len(distances) > 0:
                            best_idx = np.argmin(distances)
                            if distances[best_idx] < 0.45:
                                matched_id = known_names[best_idx]
                                if matched_id.lower() != reg_no.lower():
                                    face_is_duplicate = True
                                    duplicate_id = matched_id
            except Exception as e:
                print("Face val error:", e)

        if face_is_duplicate:
            con.close()
            return render_template("student_signup.html", error=f"Face already registered under user {duplicate_id}. Multiple accounts per face are not allowed.", preset_id=reg_no)

        cur.execute("INSERT INTO users (user_id, password, role, dob, name, year, section, branch) VALUES (?, ?, 'student', ?, ?, ?, ?, ?)",
                    (reg_no, password, dob, name, year, section, branch))
        con.commit()
        con.close()

        # Save up to 5 images
        save_dir = os.path.join(CLEAN_FACES_DIR, reg_no)
        os.makedirs(save_dir, exist_ok=True)
        for idx, b64img in enumerate(images_b64[:5]):
            try:
                header, data = b64img.split(',', 1) if ',' in b64img else ('', b64img)
                img_bytes = base64.b64decode(data)
                with open(os.path.join(save_dir, f"img_{idx+1}.jpg"), "wb") as f:
                    f.write(img_bytes)
            except Exception:
                pass

        database.init_today([reg_no])
        database.init_today_periods([reg_no])
        return render_template("reset_success.html")

    return render_template("student_signup.html")

@app.route("/hod/add-student", methods=["GET", "POST"])
def add_student():
    if session.get("role") != "hod":
        return redirect("/")
    if request.method == "POST":
        user_id  = request.form.get("user_id", "").strip().upper()
        password = request.form.get("password") or "123"
        dob      = request.form.get("dob")
        branch   = request.form.get("branch", "").strip().upper()
        year     = request.form.get("year", "").strip()
        section  = request.form.get("section", "").strip().upper()

        import os
        CLEAN_FACES_DIR = r"E:\dataset-attendance\clean_faces"
        if not os.path.exists(os.path.join(CLEAN_FACES_DIR, user_id)):
            return render_template("add_student.html",
                error=f"No image folder found for '{user_id}' in the backend dataset. "
                      "Please ensure the face images are physically added first.")

        con = get_db()
        cur = con.cursor()
        # Insert with all fields — branch is critical for HOD dept filter
        cur.execute(
            "INSERT OR IGNORE INTO users (user_id, password, role, dob, branch, year, section) "
            "VALUES (?, ?, 'student', ?, ?, ?, ?)",
            (user_id, password, dob, branch or None, year or None, section or None)
        )
        con.commit()
        con.close()
        database.init_today([user_id])
        database.init_today_periods([user_id])
        return redirect("/hod")
    return render_template("add_student.html")


# ---------- STUDENT PROFILE ----------
@app.route("/student/profile", methods=["GET", "POST"])
def student_profile():
    if session.get("role") != "student":
        return redirect("/")
    sid = session["user_id"]
    con = get_db()
    cur = con.cursor()

    if request.method == "POST":
        name = request.form.get("name")
        dob = request.form.get("dob")
        year = request.form.get("year")
        section = request.form.get("section")
        branch = request.form.get("branch")
        new_password = request.form.get("new_password")
        cur.execute("UPDATE users SET name=?, dob=?, year=?, section=?, branch=? WHERE user_id=?",
                    (name, dob, year, section, branch, sid))
        if new_password:
            cur.execute("UPDATE users SET password=? WHERE user_id=?", (new_password, sid))
        con.commit()
        con.close()
        return redirect("/student/profile?saved=1")

    cur.execute("SELECT user_id, name, dob, year, section, branch FROM users WHERE user_id=?", (sid,))
    user = cur.fetchone()
    con.close()
    saved = request.args.get("saved")
    return render_template("student_profile.html", user=user, saved=saved)


# ---------- HOD DELETE STUDENT ----------
@app.route("/hod/delete-student", methods=["POST"])
def delete_student():
    if session.get("role") != "hod":
        return redirect("/")
    student_id = request.form.get("student_id")
    con = get_db()
    cur = con.cursor()
    cur.execute("DELETE FROM users WHERE user_id=? AND role='student'", (student_id,))
    cur.execute("DELETE FROM attendance WHERE student_id=?", (student_id,))
    cur.execute("DELETE FROM period_attendance WHERE student_id=?", (student_id,))
    con.commit()
    con.close()
    return redirect("/hod")

@app.route("/api/student-images/<user_id>")
def api_student_images(user_id):
    if session.get("role") != "hod":
        return jsonify({"error": "unauthorized"}), 401
    import os
    CLEAN_FACES_DIR = r"E:\dataset-attendance\clean_faces"
    folder = os.path.join(CLEAN_FACES_DIR, user_id)
    images = []
    if os.path.exists(folder):
        images = [f"/dataset/images/{user_id}/{img}" for img in os.listdir(folder) if img.lower().endswith(('.png', '.jpg', '.jpeg'))]
    return jsonify({"images": images})

@app.route("/dataset/images/<user_id>/<filename>")
def serve_student_image(user_id, filename):
    if session.get("role") != "hod":
        return "Unauthorized", 401
    import os
    from flask import send_from_directory
    CLEAN_FACES_DIR = r"E:\dataset-attendance\clean_faces"
    folder = os.path.join(CLEAN_FACES_DIR, user_id)
    return send_from_directory(folder, filename)


# ---------- TOGGLE ATTENDANCE ----------
@app.route("/api/toggle-attendance", methods=["POST"])
def api_toggle_attendance():
    if session.get("role") != "hod":
        return jsonify({"error": "unauthorized"}), 401
    data = request.get_json()
    student_id = data.get("student_id")
    date_str = data.get("date")
    period = data.get("period")
    new_status = data.get("status", "present")  # 'present', 'absent'
    time_now = dbmod.now_ist().strftime("%H:%M:%S")
    status_map = {'present': 1.0, 'absent': 0.0}
    value = status_map.get(new_status, 0.0)
    con = get_db()
    cur = con.cursor()
    # Update period_attendance
    if period:
        cur.execute("UPDATE period_attendance SET status=?, value=?, first_seen_time=? WHERE student_id=? AND date=? AND period=?",
                    (new_status, value, time_now if new_status != 'absent' else None, student_id, date_str, period))
    # Also update legacy table
    legacy_present = 1 if new_status == 'present' else 0
    cur.execute("UPDATE attendance SET present=?, first_seen_time=? WHERE student_id=? AND date=?",
                (legacy_present, time_now if legacy_present else None, student_id, date_str))
    con.commit()
    con.close()
    return jsonify({"ok": True, "status": new_status, "value": value})


# ---------- EXPORT EXCEL ----------
@app.route("/hod/export-excel")
def export_excel():
    if session.get("role") != "hod":
        return redirect("/")
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        return "openpyxl not installed. Run: pip install openpyxl", 500

    time_filter = request.args.get("filter", "all")
    year_filter = request.args.get("year")
    branch_filter = request.args.get("branch")
    section_filter = request.args.get("section")

    con = get_db()
    cur = con.cursor()
    query = "SELECT pa.student_id, pa.date, pa.period, pa.status, pa.value, pa.first_seen_time FROM period_attendance pa LEFT JOIN users u ON u.user_id = pa.student_id"
    where = []
    params = []
    if time_filter != "all":
        days = {"today": 0, "7days": 7, "15days": 15, "30days": 30}.get(time_filter, 7)
        start_date = (dbmod.now_ist() - timedelta(days=days)).strftime("%Y-%m-%d")
        where.append("pa.date >= ?")
        params.append(start_date)
    if year_filter:
        where.append("(u.year = ? OR (CASE WHEN substr(pa.student_id,1,2) IN ('21','22') THEN '4' WHEN substr(pa.student_id,1,2) = '23' THEN '3' WHEN substr(pa.student_id,1,2) = '24' THEN '2' WHEN substr(pa.student_id,1,2) = '25' THEN '1' ELSE NULL END) = ?)")
        params.extend([year_filter, year_filter])
    if branch_filter:
        where.append("(lower(u.branch) = lower(?) OR lower(pa.student_id) LIKE '%' || lower(?) || '%')")
        params.extend([branch_filter, branch_filter])
    if section_filter:
        where.append("lower(u.section) = lower(?)")
        params.append(section_filter)
    if where:
        query += " WHERE " + " AND ".join(where)
    query += " ORDER BY pa.date DESC, pa.student_id ASC, pa.period ASC"
    cur.execute(query, params)
    rows = cur.fetchall()
    con.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Period Attendance Report"

    # Header styling
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    headers = ["Student ID", "Date", "Period", "Status", "Value", "Time In"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    present_fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
    absent_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")

    for r_idx, (sid, dt, period, status, value, tm) in enumerate(rows, 2):
        ws.cell(row=r_idx, column=1, value=sid).border = thin_border
        ws.cell(row=r_idx, column=2, value=dt).border = thin_border
        ws.cell(row=r_idx, column=3, value=f"P{period}").border = thin_border
        status_cell = ws.cell(row=r_idx, column=4, value=status.capitalize())
        status_cell.border = thin_border
        if status == 'present':
            status_cell.fill = present_fill
        else:
            status_cell.fill = absent_fill
        status_cell.alignment = Alignment(horizontal='center')
        ws.cell(row=r_idx, column=5, value=value).border = thin_border
        ws.cell(row=r_idx, column=6, value=tm or "--").border = thin_border

    for col in range(1, 7):
        ws.column_dimensions[chr(64+col)].width = 18

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"period_attendance_report_{dbmod.now_ist().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(buf, download_name=fname, as_attachment=True,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# ---------- HOD STATS API ----------
@app.route("/api/hod-stats")
def api_hod_stats():
    if session.get("role") != "hod":
        return jsonify({"error": "unauthorized"}), 401
    con = get_db()
    cur = con.cursor()
    # Total students
    cur.execute("SELECT COUNT(*) FROM users WHERE role='student'")
    total_students = cur.fetchone()[0]
    # Today's period attendance
    today_str = dbmod.now_ist().strftime("%Y-%m-%d")
    cur.execute("SELECT COUNT(*) FROM period_attendance WHERE date=? AND status='present'", (today_str,))
    present_today = cur.fetchone()[0]
    cur.execute("SELECT COUNT(*) FROM period_attendance WHERE date=?", (today_str,))
    total_today = cur.fetchone()[0]
    # Last 7 days trend
    trend = []
    for i in range(6, -1, -1):
        d = (dbmod.now_ist() - timedelta(days=i)).strftime("%Y-%m-%d")
        cur.execute("SELECT COUNT(*) FROM period_attendance WHERE date=? AND status='present'", (d,))
        p = cur.fetchone()[0]
        cur.execute("SELECT COUNT(*) FROM period_attendance WHERE date=?", (d,))
        t = cur.fetchone()[0]
        trend.append({"date": d, "present": p, "total": t})
    con.close()
    return jsonify({
        "total_students": total_students,
        "present_today": present_today,
        "total_today": total_today,
        "trend": trend
    })
if __name__ == '__main__':
    # Listen on all network interfaces (0.0.0.0) so mobiles and other laptops on the same Wi-Fi can access the portal via this computer's IP address.
    app.run(host="0.0.0.0", port=5000, debug=True)
